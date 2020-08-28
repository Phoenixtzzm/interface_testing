# _*_ coding:utf-8 _*_
__author__ = 'Tanz'
__date__ = '2018/6/22 15:09'

import os, sys
sys.path.append(os.pardir)
import openpyxl


class HandleExcel(object):
    def load_excel(self, excel_path):
        open_excel = openpyxl.load_workbook(excel_path)
        return open_excel

    # 获取sheetdata对象
    def get_sheet_value(self, index=None, excel_path=None):
        if index == None:
            self.index = 0
        else:
            self.index = index
        if excel_path == None:
            default_excel_path = os.path.abspath('..') + '/case/case.xlsx'
            self.excel_path = default_excel_path
        else:
            self.excel_path = excel_path
        # sheet_value = excel_object[sheetnames]
        sheet_data = self.load_excel(self.excel_path)[self.load_excel(self.excel_path).sheetnames[self.index]]
        return sheet_data

    def get_cell_value(self, row, col):
        cell_value = self.get_sheet_value().cell(row, col).value
        return cell_value

    def get_rows(self):
        rows = self.get_sheet_value().max_row
        return rows

    def get_row_value(self, row):
        row_value = []
        row_object_list = self.get_sheet_value()[row]

        for i in row_object_list:
            row_value.append(i.value)
        return row_value

    def excel_write_data(self, row, col, value, excel_path=None):
        if excel_path == None:
            default_excel_path = os.path.abspath('..') + '/case/case.xlsx'
            self.excel_path = default_excel_path
        else:
            self.excel_path = excel_path
        wb = self.load_excel(self.excel_path)
        # 指定写入某个sheet，默认0
        wb.active = 0
        ws = wb.active
        ws.cell(row=row, column=col, value=value)
        wb.save(self.excel_path)

    def get_col_data(self, key=None):
        '''
        获取某一列的数据
        '''
        col_list = []
        if key == None:
            self.key = 'A'
        else:
            self.key = key
        for i in self.get_sheet_value()[self.key]:
            col_list.append(i.value)
        return col_list

    # 获取行号
    def get_row_number(self, case_id):
        num = 1
        col_data = self.get_col_data(key='A')
        for col_data in col_data:
            if case_id == col_data:
                return num
            num = num + 1

        return num

    # 获取excel所有的数据
    def get_excel_data(self):
        data_list = []
        for i in range(self.get_rows()):
            row_data = self.get_row_value(i+2)
            data_list.append(row_data)
        return data_list
if __name__ == '__main__':
    a = HandleExcel()
    d = a.get_cell_value(3,9)
    print(d)
